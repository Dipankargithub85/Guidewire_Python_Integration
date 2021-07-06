# -*- coding: utf-8 -*-
"""
Created on Thu Feb 21 08:45:21 2019

@author: Dipankar
"""


#import filecmp
#from filecmp import dircmp
import shutil
import os
import xlsxwriter
import openpyxl
from difflib import SequenceMatcher
from openpyxl import Workbook
from openpyxl import load_workbook
from datetime import datetime
#from openpyxl.styles.borders import Border, Side
from openpyxl.styles import NamedStyle, Font, Border, Side
from openpyxl.styles import PatternFill
import pandas as pd
from copy import copy
#import xlsxwriter


def getdateString():
    now = datetime.now() 
    daymon = now.strftime("%d-%b")
    return daymon

def getyear():
    now = datetime.now() 
    yr = now.strftime("%Y")
    return yr
 
def getdirname(d1,d2,name):
    dlist1=list()
    dlist2=list()
    dlistsr1=list()
    dlistsr2=list()
    totcntlist=list()
    changelist=list()
    mninetylist=list()
    addlist=list()
    matchlist=list()
#    print(d1)
#    print(d2)
    filevel=0
    
    for (dirpath, dirnames, filenames) in os.walk(d1):
        for d in dirnames:
            dlist1.append(os.path.join(dirpath,d))
            dlistsr1.append(d)
            filevel = filevel + 1
            
        if filevel > 0:
            break

    filevel =0
    for (dirpath, dirnames, filenames) in os.walk(d2):
        for d in dirnames:
            dlist2.append(os.path.join(dirpath,d))
            dlistsr2.append(d)
            filevel = filevel +1 
            
        if filevel > 0:
            break
    
    

    
    skipcnt=0
    cnt=0
    flag=False
    for i in range(0,len(dlist1)):
        emptyTuple = ()
        skipcnt=0
        totflcnt=0
        #print(dlist1[i])
        for j in range(0,len(dlist2)):
            
            if cnt != skipcnt:
                skipcnt = skipcnt+1
                flag = True
                continue
                
            if dlistsr1[i] == dlistsr2[j]:
                
                emptyTuple = getChangeCount(dlist1[i],dlist2[j])
                cnt= cnt + 1
                totcntlist.append(emptyTuple[0])
                matchlist.append(emptyTuple[1])
                changelist.append(emptyTuple[2])
                mninetylist.append(emptyTuple[3])
                addlist.append(emptyTuple[4])
                flag = False
                
                break
            else:
                totflcnt= getFileCnt(dlist1[i])
               # changelist.append(totflcnt)
                addlist.append(totflcnt)
                flag = False
                break
        
        if flag:
            totflcnt= getFileCnt(dlist1[i])
            #changelist.append(totflcnt)
            addlist.append(totflcnt)
            


        
    
    totpcffilecnt=0
    changepcffilecnt=0  
    cngnity=0    
    addfile=0 
    matchfl=0    
    addWeightage=0
    chgWeightage=0
    pcfAlginement=0
    for  i in range(0,len(totcntlist)):
        totpcffilecnt = totpcffilecnt + totcntlist[i]
        
        
    for  b in range(0,len(matchlist)):
        matchfl = matchfl + matchlist[b]
        
    
    
    for  i in range(0,len(changelist)):
        changepcffilecnt = changepcffilecnt + changelist[i]
        
    
    for k in range (0,len(mninetylist)):
        cngnity = cngnity + mninetylist[k]
        
    for j in range (0,len(addlist)):
        addfile = addfile + addlist[j]
        
    
    
    addper = (addfile/totpcffilecnt)
    changeper = (changepcffilecnt/totpcffilecnt)
    if name =='PCF':
        addWeightage=1
        chgWeightage=0.5
        pcfAlginement = (addper*addWeightage) + (changeper*chgWeightage)
        
    if name =='Entity(editable)':
        addWeightage=1
        chgWeightage=1
        pcfAlginement = (addper*addWeightage) + (changeper*chgWeightage)
        
        
    
    if name =='Type':
        addWeightage=0
        chgWeightage=0
        pcfAlginement = (addper*addWeightage) + (changeper*chgWeightage)
        
        
    if name =='Rule':
        addWeightage=0.2
        chgWeightage=0
        pcfAlginement = (addper*addWeightage) + (changeper*chgWeightage)
        
        
    if name =='Intake':
        addWeightage=1
        chgWeightage=0.5
        pcfAlginement = (addper*addWeightage) + (changeper*chgWeightage)
        
    
    percetange=  (changepcffilecnt + addfile) * 100 / totpcffilecnt
    statdirtp =()
    #statdirtp=(name,str(totpcffilecnt),str(matchfl),str(cngnity),str(changepcffilecnt),str(addfile),str(addper),str(addWeightage),str(changeper),str(chgWeightage),str(pcfAlginement))
    statdirtp=(name,str(totpcffilecnt),str(matchfl),str(cngnity),str(changepcffilecnt),str(addfile),addper,addWeightage,changeper,chgWeightage,pcfAlginement)
    return statdirtp       

def getEntityPer(ed1,ed2,name):
    newentytp=()
    newentytp= getcntfiles(ed1,ed2)
    addWeightage=0
    chgWeightage=0
    pcfAlginement=0
   # percetange=  (newentytp[2] + newentytp[4])  * 100 / newentytp[0]
    addper = (newentytp[4]/newentytp[0])
    changeper = (newentytp[2]/newentytp[0])
           
    if name =='Entity(editable)':
        addWeightage=1
        chgWeightage=1
        pcfAlginement = (addper*addWeightage) + (changeper*chgWeightage)
        
        
    
        
   # statfltp=(name,str(newentytp[0]),str(newentytp[1]),str(newentytp[3]),str(newentytp[2]),str(newentytp[4]),str(addper),str(addWeightage),str(changeper),str(chgWeightage),str(pcfAlginement))
    statfltp=(name,str(newentytp[0]),str(newentytp[1]),str(newentytp[3]),str(newentytp[2]),str(newentytp[4]),addper,addWeightage,changeper,chgWeightage,pcfAlginement)
    return statfltp
  #  return percetange
 
    
def dirandFilecompre(fn1,fn2,name):
    fnoltp= ()
    newentytp=()
    newentytp= getChangeCount(fn1,fn2)
    addWeightage=0
    chgWeightage=0
    pcfAlginement=0
    addper = (newentytp[4]/newentytp[0])
    changeper = (newentytp[2]/newentytp[0])
           
    if name =='Intake':
        addWeightage=1
        chgWeightage=0.5
        pcfAlginement = (addper*addWeightage) + (changeper*chgWeightage)

    fnoltp=(name,str(newentytp[0]),str(newentytp[1]),str(newentytp[3]),str(newentytp[2]),str(newentytp[4]),addper,addWeightage,changeper,chgWeightage,pcfAlginement)
    return fnoltp  
    
def writeXcel(StatList):
    
    GetYear = getyear();
    SheetName = getdateString();
    fileName = 'GW_StatFile_' + GetYear+'-'+getdateString()+'.xlsx'
    filePath = 'C:\\TestTmp\\' + fileName

    exists = os.path.isfile(filePath)
    #flag=False
    if not exists:
        workbook  = xlsxwriter.Workbook(filePath)

        worksheet1 = workbook.add_worksheet(SheetName)
        merge_format = workbook.add_format({
                'bold': 1,
                'border': 1,
                'align': 'center',
                'valign': 'vcenter'
                })
        row = 0
        col = 0
        val="1"
        worksheet1.set_column(0, 6, 20) 
        if len(StatList) > 0:
            getheader(workbook,worksheet1,val)
            cell_formatt = workbook.add_format()
            cell_formatt.set_border()
        for adfl in StatList:
            row +=1
            if adfl[0] != 'Intake':
                worksheet1.write_string(row,col,adfl[0],cell_formatt)
                worksheet1.write_string(row,col+1,adfl[1],cell_formatt)
                worksheet1.write_string(row,col+2,adfl[2],cell_formatt)
                worksheet1.write_string(row,col+3,adfl[3],cell_formatt)
                worksheet1.write_string(row,col+4,adfl[4],cell_formatt)
                worksheet1.write_string(row,col+5,adfl[5],cell_formatt)
                cell_formatt = workbook.add_format({'num_format': '0.00%'})
                cell_formatt.set_border()
                worksheet1.write_number(row,col+6,adfl[6],cell_formatt)
              #  worksheet1.write_number(row,col+6,adfl[6],cell_formatt)
                cell_formatt = workbook.add_format()
                cell_formatt.set_border()
                worksheet1.write_number(row,col+7,adfl[7],cell_formatt)
                cell_formatt = workbook.add_format({'num_format': '0.00%'})
                cell_formatt.set_border()
                worksheet1.write_number(row,col+8,adfl[8],cell_formatt)
               # worksheet1.write_number(row,col+8,adfl[8],cell_formatt)
                cell_formatt = workbook.add_format()
                cell_formatt.set_border()
                worksheet1.write_number(row,col+9,adfl[9],cell_formatt)
                cell_formatt = workbook.add_format({'num_format': '0.00%'})
                cell_formatt.set_border()
                worksheet1.write_number(row,col+10,adfl[10],cell_formatt)
                #worksheet1.write_number(row,col+10,adfl[10],cell_formatt)
                
        cell_formatt = workbook.add_format()
        cell_formatt.set_border()
        worksheet1.merge_range('A6:F6', 'Total', merge_format)
        cell_formatt = workbook.add_format({'num_format': '0.00%'})
        cell_formatt.set_border()
        worksheet1.write_formula('G6','=SUMPRODUCT(G2:G5,H2:H5)/SUM(H2:H5)',cell_formatt)      
        worksheet1.write_formula('I6','=SUMPRODUCT(I2:I5,J2:J5)/SUM(J2:J5)',cell_formatt)
        cell_formatt = workbook.add_format()
        cell_formatt.set_border()
        worksheet1.merge_range('F8:G8', 'Total Deviation', merge_format)
        cell_formatt = workbook.add_format({'num_format': '0.00%'})
        cell_formatt.set_border()
        worksheet1.write_formula('K8','=SUM(K2:K5)/3',cell_formatt)
        row=10
        val="12"
        cell_formatt = workbook.add_format()
        cell_formatt.set_border()
        worksheet1.write_string(row,col,'Intake Specific',cell_formatt)
        getheader(workbook,worksheet1,val)
        row = 12
        for adfl in StatList:
            
            if adfl[0] == 'Intake':
                worksheet1.write_string(row,col,'PCF',cell_formatt)
                worksheet1.write_string(row,col+1,adfl[1],cell_formatt)
                worksheet1.write_string(row,col+2,adfl[2],cell_formatt)
                worksheet1.write_string(row,col+3,adfl[3],cell_formatt)
                worksheet1.write_string(row,col+4,adfl[4],cell_formatt)
                worksheet1.write_string(row,col+5,adfl[5],cell_formatt)
                cell_formatt = workbook.add_format({'num_format': '0.00%'})
                cell_formatt.set_border()
                worksheet1.write_number(row,col+6,adfl[6],cell_formatt)
                cell_formatt = workbook.add_format()
                cell_formatt.set_border()
                worksheet1.write_number(row,col+7,adfl[7],cell_formatt)
                cell_formatt = workbook.add_format({'num_format': '0.00%'})
                cell_formatt.set_border()
                worksheet1.write_number(row,col+8,adfl[8],cell_formatt)
                cell_formatt = workbook.add_format()
                cell_formatt.set_border()
                worksheet1.write_number(row,col+9,adfl[9],cell_formatt)
                cell_formatt = workbook.add_format({'num_format': '0.00%'})
                cell_formatt.set_border()
                worksheet1.write_number(row,col+10,adfl[10],cell_formatt)
         
            
        cell_formatt = workbook.add_format()
        cell_formatt.set_border()    
        worksheet1.merge_range('C14:F14', 'Total', merge_format)
        cell_formatt = workbook.add_format({'num_format': '0.00%'})
        cell_formatt.set_border()
        worksheet1.write_formula('G14','=SUMPRODUCT(G13:G13,H13:H13)/SUM(H13:H13)',cell_formatt)      
        worksheet1.write_formula('I14','=SUMPRODUCT(I13:I13,J13:J13)/SUM(J13:J13)',cell_formatt)
        cell_formatt = workbook.add_format()
        cell_formatt.set_border()
        worksheet1.merge_range('F16:G16', 'Total Deviation', merge_format)
        cell_formatt = workbook.add_format({'num_format': '0.00%'})
        cell_formatt.set_border()
        worksheet1.write_formula('K16','=SUM(K13)',cell_formatt)
 
        workbook.close()
        

    else:
        print("Please Check the File Name")
        
    
    copyXcel(filePath,SheetName)
  
def buildTreandData(updatetpl,SheetName):
    filename = updatetpl[0]
    alngvalue = updatetpl[1]
    intakevalue = updatetpl[2]
    openxl = openpyxl.load_workbook(filename)
    shtname = openxl.get_sheet_by_name('Trend')
    totsheetcnt = openxl.sheetnames
    A1=shtname.cell(row = totsheetcnt+1, column = 1)
    A1.value =SheetName
    B1=shtname.cell(row = totsheetcnt+1, column = 2)
    B1.value =alngvalue
    C1=shtname.cell(row = totsheetcnt+1, column = 3)
    C1.value =intakevalue
    
    
    

def copyXcel(filePath,SheetName):
    wbnew  = openpyxl.load_workbook(filePath)
  #  newSheet = wbnew.get_sheet_by_name(SheetName);
    updateFilepath='C:\\TestTmp\\OOTB_Report_2019.xlsx'
    newSheet = wbnew.worksheets[0]
    destfile = openpyxl.load_workbook(updateFilepath)
    totshtname = list()
    totshtname=destfile.sheetnames
    totshtcnt = len(totshtname) -1 # index start with zero but count start with 1

    if totshtcnt < 54:# one Year Data
        destfileSht = destfile.create_sheet(index = totshtcnt , title = SheetName) 
    
        for row in newSheet:
            for cell in row:
                new_cell = destfileSht.cell(row=cell.row, column=cell.col_idx,
                                            value= cell.value)
               

        
                if cell.has_style:
                    new_cell.font = copy(cell.font)
                    new_cell.border = copy(cell.border)
                    new_cell.fill = copy(cell.fill)
                    new_cell.number_format = copy(cell.number_format)
           # new_cell.protection = copy(cell.protection)
                    new_cell.alignment = copy(cell.alignment)
                  #  new_cell.f
        destfile.save('C:\\TestTmp\\OOTB_Report_2019.xlsx')
        #localtupl=(updateFilepath,algnvlaue,Intakevlaue)
     ##    return localtupl
    else:
        print("Please rename the spread sheet")
       # return ""    
   
    
 
 
    
def getheader(wb,ws,val):
    
    cell_format = wb.add_format()
    cell_format.set_text_wrap()
    cell_format.set_pattern(1)  # This is optional when using a solid fill.
    cell_format.set_bg_color('#FFCC99')
    cell_format.set_bold()
   # cell_format.set_font_color('red')
    cell_format.set_border()
    ws.write("A"+val, 'Object Type', cell_format)
    ws.write("B"+val, 'Total Base File Count OOTB ', cell_format)
    ws.write("C"+val, 'Match count', cell_format)
    ws.write("D"+val, 'Less than 90% change', cell_format)
    ws.write("E"+val, 'Change count', cell_format)
    ws.write("F"+val, 'Add file Count', cell_format)
    ws.write("G"+val, '% Added', cell_format)
    ws.write("H"+val, 'Weightage', cell_format)
    ws.write("I"+val, '% Change', cell_format)
    ws.write("J"+val, 'Weightage', cell_format)
    ws.write("K"+val, 'Overall Alginement%', cell_format)
    
 
    
def getChangeCount(dr1,dr2):
    
    emptyTp =()
    emptydir = ()
    listOfFiles1 = list()
    for (dirpath, dirnames, filenames) in os.walk(dr1):
        listOfFiles1 += [os.path.join(dirpath, file) for file in filenames]
     
    #print(listOfFiles1)
     
    listOfFiles2 = list()
    for (dirpath, dirnames, filenames) in os.walk(dr2):
        listOfFiles2 += [os.path.join(dirpath, file) for file in filenames]
        
    
    
    emptydir = createandCopyDir(listOfFiles1,listOfFiles2)
    dirval1 = emptydir[0]
    dirval2 = emptydir[1]
    emptyTp = getcntfiles(dirval1,dirval2) 
    shutil.rmtree(emptydir[0], ignore_errors=True)
    shutil.rmtree(emptydir[1], ignore_errors=True)
    return emptyTp
    
        

def createandCopyDir(l1,l2):
    
    dirName1 = 'C:\\testtemp\\tempDiror'
    dirName2 = 'C:\\testtemp\\tempDirmod'
    emptydir = ()
    try:
        # Create target Directory
        
        
        os.mkdir(dirName1)

    except FileExistsError:
        print("Directory " , dirName1 ,  "already exists")
        
    try:
        os.mkdir(dirName2)

    except FileExistsError:
        print("Directory " , dirName1 ,  "already exists")
     
        
    for f in l1:
        shutil.copy(f, dirName1)
    for ff in l2:
        shutil.copy(ff, dirName2)
    
    emptydir = (dirName1,dirName2)

    return emptydir


def getcntfiles(dr1,dr2):
    emptyTp = ()
    
    d1_contents = list(os.listdir(dr1))
    d2_contents = list(os.listdir(dr2))

    mcnt=0
    miscnt = 0
    flag = False
    mninety=0
    addfile=0
    totfile=0
    changefl = 0
    for i in range(0,len(d1_contents)):
        skipcnt=0

        for j in range(0,len(d2_contents)):
            
            if totfile != skipcnt:
                skipcnt = skipcnt+1
                flag = True
                continue
 
            

            if d1_contents[i] == d2_contents[j]:
#                print(d1_contents[i])
#                print(d2_contents[j])
                totfile = totfile + 1
                filepathor= open(os.path.join(dr1,d1_contents[i]),encoding='utf-8').read()
                filepathmod =open(os.path.join(dr2,d2_contents[j]),encoding='utf-8').read()
                m = SequenceMatcher(None, filepathor, filepathmod)
               # abc = m.ratio()
                if str(m.ratio()*100)[:1] == '1':
                    mcnt= mcnt + 1
              #  totcnt = totcnt +1
                    flag = False
                    break
                elif str(m.ratio()*100)[:2] >= '90':
                    #mcnt= mcnt + 1
                    mninety = mninety + 1
                    flag = False
                    break
                    
                else:
                    #mcnt= mcnt + 1
                    changefl= changefl+1
                    flag = False
                    break
            else:
                flag = False
                miscnt= miscnt+1
                addfile = addfile + 1
               # changelist.append(d1_contents[i])
                break
        
        
        
        if flag:
            miscnt= miscnt+1
            addfile = addfile + 1
  
   
    emptyTp = (totfile,mcnt,changefl,mninety,addfile)
    
    return emptyTp

        

def getFileCnt(dr1):
    listOfFiles1 = list()
    for (dirpath, dirnames, filenames) in os.walk(dr1):
        
        listOfFiles1 += [os.path.join(file) for file in filenames]
        return len(listOfFiles1)
    

def main():
    print("Started")
    mainDir = 'C:\\testtemp'
   
    if not os.path.exists(mainDir):
        os.mkdir(mainDir)
    else:
        print("Directory " , mainDir ,  "already exists")
    
    StatList=list()
    temptp=()    
   
    #PCF fiels
    pcfmod='C:\\GW\\workspace\\v10_conversion\\ClaimCenter\\modules\\configuration\\config\\web\\pcf'
    #pcforg='C:\\BaseGW\\ClaimCenter\\modules\\configuration\\config\\web\\pcf'
    pcforg='C:\\BaseGW\\ClaimCenter\\modules\\base\\config\\web\\pcf'
    name='PCF'
    temptp = getdirname(pcfmod,pcforg,name)
    StatList.append(temptp)
    temptp=()
    #Entity fiels
    entmod='C:\\GW\\workspace\\v10_conversion\\ClaimCenter\\modules\\configuration\\config\\extensions\\entity'
    entorg='C:\\BaseGW\\ClaimCenter\\modules\\base\\config\\extensions\\entity'
   # entorg='C:\\BaseGW\\ClaimCenter\\modules\\configuration\\config\\extensions\\entity'
    name='Entity(editable)'
    temptp = getEntityPer(entmod,entorg,name)
    StatList.append(temptp)
    temptp=()
    #TypeList fiels
    tlistmod='C:\\GW\\workspace\\v10_conversion\\ClaimCenter\\modules\\configuration\\config\\extensions\\typelist'
    tlistorg='C:\\BaseGW\\ClaimCenter\\modules\\base\\config\\extensions\\typelist'
   # tlistorg='C:\\BaseGW\\ClaimCenter\\modules\\configuration\\config\\extensions\\typelist'
    name='Type'
    temptp = getEntityPer(tlistmod,tlistorg,name)
    StatList.append(temptp)
    temptp=()
    #Rules fiels
    rulemod='C:\\GW\\workspace\\v10_conversion\\ClaimCenter\\modules\\configuration\\config\\rules'
    ruleOrg='C:\\BaseGW\\ClaimCenter\\modules\\base\\config\\rules'
    #ruleOrg='C:\\BaseGW\\ClaimCenter\\modules\\configuration\\config\\rules'
    name='Rule'
    temptp = getdirname(rulemod,ruleOrg,name)
    StatList.append(temptp)
    temptp=()
    
    #Intake fiels
    intakemod='C:\\GW\workspace\\v10_conversion\\ClaimCenter\\modules\\configuration\\config\\web\\pcf\\claim\\FNOL'
    intakeorg='C:\\BaseGW\\ClaimCenter\\modules\\base\\config\\web\\pcf\\claim\\FNOL'
  #  intakeorg='C:\\BaseGW\\ClaimCenter\\modules\\configuration\\config\\web\\pcf\\claim\\FNOL'
    name='Intake'
    temptp = dirandFilecompre(intakemod,intakeorg,name)
    StatList.append(temptp)
    writeXcel(StatList)
    print("End")

if __name__== "__main__":
  main()